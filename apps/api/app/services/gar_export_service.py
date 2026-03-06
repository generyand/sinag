# GAR Export Service
# Generates Excel and PDF exports for GAR reports

import io
import logging
import os

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from app.schemas.bbi import AssessmentBBIComplianceResponse
from app.schemas.gar import (
    GARChecklistItem,
    GARGovernanceArea,
    GARIndicator,
    GARResponse,
)

# Logo directory for PDF exports
_LOGOS_DIR = os.path.join(os.path.dirname(__file__), "..", "static", "logos")


class GARExportService:
    """Service for exporting GAR reports to Excel and PDF."""

    def __init__(self):
        self.logger = logging.getLogger(__name__)

        # Define colors matching the official GAR format
        self.colors = {
            "met": "90EE90",  # Light green
            "considered": "FFFF00",  # Yellow
            "unmet": "FF6B6B",  # Light red
            "header_bg": "90EE90",  # Light green for area headers
            "overall_bg": "90EE90",  # Light green for overall result
            # BBI compliance colors (DILG MC 2024-417)
            "bbi_highly_functional": "90EE90",  # Light green (75%+)
            "bbi_moderately_functional": "FFD700",  # Gold/amber (50-74%)
            "bbi_low_functional": "FF6B6B",  # Light red (<50%)
            "bbi_header_bg": "B0C4DE",  # Light steel blue
        }

        # Define border style
        self.thin_border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )

    def generate_excel(self, gar_data: GARResponse) -> io.BytesIO:
        """
        Generate Excel file for GAR report.

        Args:
            gar_data: GAR response data

        Returns:
            BytesIO buffer containing the Excel file
        """
        wb = Workbook()
        ws = wb.active
        ws.title = "GAR Report"

        # Set column widths
        ws.column_dimensions["A"].width = 80  # Indicators column
        ws.column_dimensions["B"].width = 15  # Result column

        current_row = 1

        # Add header
        current_row = self._add_header(ws, gar_data, current_row)

        # Add each governance area
        for area in gar_data.governance_areas:
            current_row = self._add_governance_area(ws, area, current_row)
            current_row += 1  # Add spacing between areas

        # Add summary table if multiple areas
        if len(gar_data.governance_areas) > 1:
            current_row = self._add_summary_table(ws, gar_data, current_row)

        # Add BBI compliance section (DILG MC 2024-417)
        if gar_data.bbi_compliance and gar_data.bbi_compliance.bbi_results:
            current_row = self._add_bbi_compliance_table(ws, gar_data.bbi_compliance, current_row)

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer

    def _add_header(self, ws, gar_data: GARResponse, start_row: int) -> int:
        """Add GAR header with barangay info."""
        row = start_row

        # Title
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = gar_data.cycle_year
        cell.font = Font(bold=True, size=14)
        cell.alignment = Alignment(horizontal="center")
        row += 1

        # Subtitle
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = "BARANGAY GOVERNANCE ASSESSMENT REPORT"
        cell.font = Font(bold=True, size=12)
        cell.alignment = Alignment(horizontal="center")
        row += 1

        # Barangay info
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = (
            f"Barangay {gar_data.barangay_name}, {gar_data.municipality}, {gar_data.province}"
        )
        cell.font = Font(bold=True, italic=True, size=11)
        cell.alignment = Alignment(horizontal="center")
        row += 2  # Extra spacing

        return row

    def _add_governance_area(self, ws, area: GARGovernanceArea, start_row: int) -> int:
        """Add a single governance area section to the worksheet."""
        row = start_row

        # Area header (e.g., "CGA 1: FINANCIAL ADMINISTRATION AND SUSTAINABILITY")
        area_prefix = "CGA" if area.area_type == "Core" else "EGA"
        area_num = area.area_number if area.area_type == "Core" else area.area_number - 3
        area_title = f"{area_prefix} {area_num}: {area.area_name.upper()}"

        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = area_title
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(
            start_color=self.colors["header_bg"],
            end_color=self.colors["header_bg"],
            fill_type="solid",
        )
        cell.border = self.thin_border
        ws[f"B{row}"].border = self.thin_border
        row += 1

        # Column headers
        ws[f"A{row}"].value = "INDICATORS"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = self.thin_border
        ws[f"A{row}"].alignment = Alignment(horizontal="center")

        ws[f"B{row}"].value = "RESULT"
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].border = self.thin_border
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # Add result legend in header
        ws[f"B{row - 1}"].value = "RESULT\n(met, considered, unmet)"
        ws[f"B{row - 1}"].alignment = Alignment(horizontal="center", wrap_text=True)

        # Add indicators and checklist items
        for indicator in area.indicators:
            row = self._add_indicator_row(ws, indicator, row)

            # Add checklist items under this indicator
            for item in indicator.checklist_items:
                row = self._add_checklist_row(ws, item, row)

        # Add Overall Result row
        ws[f"A{row}"].value = "OVERALL RESULT"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].fill = PatternFill(
            start_color=self.colors["overall_bg"],
            end_color=self.colors["overall_bg"],
            fill_type="solid",
        )
        ws[f"A{row}"].border = self.thin_border

        ws[f"B{row}"].border = self.thin_border
        if area.overall_result:
            result_color = (
                self.colors["met"] if area.overall_result == "Passed" else self.colors["unmet"]
            )
            ws[f"B{row}"].fill = PatternFill(
                start_color=result_color, end_color=result_color, fill_type="solid"
            )
        row += 1

        return row

    def _add_indicator_row(self, ws, indicator: GARIndicator, row: int) -> int:
        """Add an indicator row to the worksheet."""
        # Indicator name with code
        indent = "  " * indicator.indent_level
        cell_value = f"{indent}{indicator.indicator_code} {indicator.indicator_name}"

        ws[f"A{row}"].value = cell_value
        ws[f"A{row}"].border = self.thin_border
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        # Make header indicators bold
        if indicator.is_header:
            ws[f"A{row}"].font = Font(bold=True)
            ws[f"A{row}"].fill = PatternFill(
                start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"
            )

        # Result cell
        ws[f"B{row}"].border = self.thin_border
        if indicator.validation_status and not indicator.is_header:
            color = self._get_status_color(indicator.validation_status)
            if color:
                ws[f"B{row}"].fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

        return row + 1

    def _add_checklist_row(self, ws, item: GARChecklistItem, row: int) -> int:
        """Add a checklist item row to the worksheet."""
        # Skip info_text items that are just labels
        if item.item_type == "info_text" and not item.validation_result:
            # Still show info_text as headers/labels
            ws[f"A{row}"].value = f"    {item.label}"
            ws[f"A{row}"].border = self.thin_border
            ws[f"A{row}"].font = Font(italic=True)
            ws[f"B{row}"].border = self.thin_border
            return row + 1

        # Regular checklist item
        ws[f"A{row}"].value = f"    {item.label}"
        ws[f"A{row}"].border = self.thin_border
        ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="top")

        # Result cell with color
        ws[f"B{row}"].border = self.thin_border
        if item.validation_result:
            color = self._get_result_color(item.validation_result)
            if color:
                ws[f"B{row}"].fill = PatternFill(
                    start_color=color, end_color=color, fill_type="solid"
                )

        return row + 1

    def _add_summary_table(self, ws, gar_data: GARResponse, start_row: int) -> int:
        """Add summary table at the bottom of the report."""
        row = start_row + 1  # Add spacing

        # Core Governance Areas header
        ws[f"A{row}"].value = "Core Governance Area"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = self.thin_border

        ws[f"B{row}"].value = "Result"
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].border = self.thin_border
        row += 1

        # Core areas
        for item in gar_data.summary:
            if item.area_type == "Core":
                ws[f"A{row}"].value = item.area_name
                ws[f"A{row}"].border = self.thin_border

                ws[f"B{row}"].border = self.thin_border
                if item.result:
                    color = self.colors["met"] if item.result == "Passed" else self.colors["unmet"]
                    ws[f"B{row}"].fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )
                row += 1

        # Essential Governance Areas header
        ws[f"A{row}"].value = "Essential Governance Area"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = self.thin_border

        ws[f"B{row}"].value = "Result"
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].border = self.thin_border
        row += 1

        # Essential areas
        for item in gar_data.summary:
            if item.area_type == "Essential":
                ws[f"A{row}"].value = item.area_name
                ws[f"A{row}"].border = self.thin_border

                ws[f"B{row}"].border = self.thin_border
                if item.result:
                    color = self.colors["met"] if item.result == "Passed" else self.colors["unmet"]
                    ws[f"B{row}"].fill = PatternFill(
                        start_color=color, end_color=color, fill_type="solid"
                    )
                row += 1

        return row

    def _add_bbi_compliance_table(
        self, ws, bbi_data: AssessmentBBIComplianceResponse, start_row: int
    ) -> int:
        """Add BBI compliance table per DILG MC 2024-417."""
        row = start_row + 2  # Add spacing

        # Section title
        ws.merge_cells(f"A{row}:B{row}")
        cell = ws[f"A{row}"]
        cell.value = "BBI FUNCTIONALITY COMPLIANCE (DILG MC 2024-417)"
        cell.font = Font(bold=True, size=11)
        cell.fill = PatternFill(
            start_color=self.colors["bbi_header_bg"],
            end_color=self.colors["bbi_header_bg"],
            fill_type="solid",
        )
        cell.border = self.thin_border
        ws[f"B{row}"].border = self.thin_border
        row += 1

        # Column headers
        ws[f"A{row}"].value = "BBI"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = self.thin_border
        ws[f"A{row}"].alignment = Alignment(horizontal="center")

        ws[f"B{row}"].value = "COMPLIANCE"
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].border = self.thin_border
        ws[f"B{row}"].alignment = Alignment(horizontal="center")
        row += 1

        # BBI results
        for bbi in bbi_data.bbi_results:
            # BBI name
            ws[f"A{row}"].value = f"{bbi.bbi_abbreviation} - {bbi.bbi_name}"
            ws[f"A{row}"].border = self.thin_border
            ws[f"A{row}"].alignment = Alignment(wrap_text=True, vertical="center")

            # Compliance percentage and rating
            ws[f"B{row}"].value = f"{round(bbi.compliance_percentage)}%"
            ws[f"B{row}"].border = self.thin_border
            ws[f"B{row}"].alignment = Alignment(horizontal="center", vertical="center")

            # Color based on rating
            rating_color = self._get_bbi_rating_color(bbi.compliance_rating)
            if rating_color:
                ws[f"B{row}"].fill = PatternFill(
                    start_color=rating_color, end_color=rating_color, fill_type="solid"
                )
            row += 1

        # Average compliance row
        ws[f"A{row}"].value = "AVERAGE COMPLIANCE"
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"A{row}"].border = self.thin_border
        ws[f"A{row}"].fill = PatternFill(
            start_color=self.colors["bbi_header_bg"],
            end_color=self.colors["bbi_header_bg"],
            fill_type="solid",
        )

        avg_pct = bbi_data.summary.average_compliance_percentage
        ws[f"B{row}"].value = f"{round(avg_pct)}%"
        ws[f"B{row}"].font = Font(bold=True)
        ws[f"B{row}"].border = self.thin_border
        ws[f"B{row}"].alignment = Alignment(horizontal="center")

        # Color for average
        avg_rating = self._determine_rating_from_percentage(avg_pct)
        avg_color = self._get_bbi_rating_color(avg_rating)
        if avg_color:
            ws[f"B{row}"].fill = PatternFill(
                start_color=avg_color, end_color=avg_color, fill_type="solid"
            )
        row += 1

        # Summary counts row
        row += 1
        ws[f"A{row}"].value = (
            f"Summary: {bbi_data.summary.highly_functional_count} Highly Functional, "
            f"{bbi_data.summary.moderately_functional_count} Moderately Functional, "
            f"{bbi_data.summary.low_functional_count} Low Functional"
        )
        ws[f"A{row}"].font = Font(italic=True, size=9)
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        # Rating legend
        ws[
            f"A{row}"
        ].value = "Rating: 75%+ = Highly Functional, 50-74% = Moderately Functional, <50% = Low Functional"
        ws[f"A{row}"].font = Font(italic=True, size=9)
        ws.merge_cells(f"A{row}:B{row}")

        return row + 1

    def _get_bbi_rating_color(self, rating: str) -> str | None:
        """Get color for BBI compliance rating."""
        rating_upper = rating.upper() if rating else ""
        if "HIGHLY" in rating_upper:
            return self.colors["bbi_highly_functional"]
        elif "MODERATELY" in rating_upper:
            return self.colors["bbi_moderately_functional"]
        elif "LOW" in rating_upper:
            return self.colors["bbi_low_functional"]
        return None

    def _determine_rating_from_percentage(self, percentage: float) -> str:
        """Determine rating tier from percentage."""
        if percentage >= 75:
            return "HIGHLY_FUNCTIONAL"
        elif percentage >= 50:
            return "MODERATELY_FUNCTIONAL"
        return "LOW_FUNCTIONAL"

    def _get_status_color(self, status: str) -> str | None:
        """Get color for validation status."""
        # Normalize to uppercase for comparison (ValidationStatus enum values are PASS, FAIL, CONDITIONAL)
        normalized = status.upper() if status else ""
        if normalized == "PASS":
            return self.colors["met"]
        elif normalized == "CONDITIONAL":
            return self.colors["considered"]
        elif normalized == "FAIL":
            return self.colors["unmet"]
        return None

    def _get_result_color(self, result: str) -> str | None:
        """Get color for checklist validation result."""
        return self.colors.get(result)

    def _get_logo_path(self, name: str) -> str | None:
        """Get absolute path to a logo file, or None if not found."""
        path = os.path.join(_LOGOS_DIR, name)
        return path if os.path.exists(path) else None

    def generate_pdf(self, gar_data: GARResponse) -> io.BytesIO:
        """
        Generate PDF file for GAR report with official DILG letterhead.

        Portrait A4 orientation with:
        - DILG + Bagong Pilipinas logos and government letterhead in header
        - Result column with color legend (met/considered/unmet)
        - Text labels in result cells for grayscale readability
        - MLGRC + ISO 9001:2015 logos and address in footer
        - Page breaks between governance areas
        - Page numbers
        """
        try:
            from reportlab.lib import colors
            from reportlab.lib.pagesizes import A4
            from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
            from reportlab.lib.units import inch
            from reportlab.platypus import (
                Image,
                PageBreak,
                Paragraph,
                SimpleDocTemplate,
                Spacer,
                Table,
                TableStyle,
            )

            buffer = io.BytesIO()
            page_w, page_h = A4  # 595 x 842 points

            # Logo paths
            dilg_logo = self._get_logo_path("dilg.png")
            bp_logo = self._get_logo_path("bagong_pilipinas.png")
            mlgrc_logo = self._get_logo_path("mlgrc.png")
            iso_logo = self._get_logo_path("iso_9001_2015.png")

            # Footer callback: MLGRC + ISO logos and address on every page
            def _draw_footer(canvas, doc):
                canvas.saveState()
                footer_y = 0.35 * inch
                center_x = page_w / 2.0

                # Footer text
                canvas.setFont("Helvetica-Bold", 7)
                canvas.drawCentredString(center_x, footer_y + 42, "\u201cMatino, Mahusay at Maaasahan\u201d")
                canvas.setFont("Helvetica", 6.5)
                canvas.drawCentredString(center_x, footer_y + 32, "MLGRC, Liga Bldg., Municipal Hall Compound")
                canvas.drawCentredString(center_x, footer_y + 23, "Prk. 9 Brgy. Poblacion, Sulop, Davao del Sur")
                canvas.setFont("Helvetica", 6.5)
                canvas.setFillColor(colors.blue)
                canvas.drawCentredString(center_x, footer_y + 14, "sulopdilg2021@gmail.com")
                canvas.setFillColor(colors.black)

                # Page number
                canvas.setFont("Helvetica", 7)
                canvas.drawCentredString(center_x, footer_y, f"Page {canvas.getPageNumber()}")

                # MLGRC logo (left of footer text)
                if mlgrc_logo:
                    try:
                        canvas.drawImage(mlgrc_logo, 0.6 * inch, footer_y + 12, width=40, height=40, preserveAspectRatio=True, mask="auto")
                    except Exception:
                        pass

                # ISO logo (right of footer text)
                if iso_logo:
                    try:
                        canvas.drawImage(iso_logo, page_w - 0.6 * inch - 40, footer_y + 12, width=40, height=40, preserveAspectRatio=True, mask="auto")
                    except Exception:
                        pass

                canvas.restoreState()

            doc = SimpleDocTemplate(
                buffer,
                pagesize=A4,
                topMargin=0.5 * inch,
                bottomMargin=1.0 * inch,
                leftMargin=0.5 * inch,
                rightMargin=0.5 * inch,
            )
            elements = []
            styles = getSampleStyleSheet()

            # Styles
            header_text_style = ParagraphStyle(
                "HeaderText", parent=styles["Normal"], fontSize=8, alignment=1, leading=10,
            )
            header_bold_style = ParagraphStyle(
                "HeaderBold", parent=styles["Normal"], fontSize=9, alignment=1, leading=11, fontName="Helvetica-Bold",
            )
            title_style = ParagraphStyle(
                "GARTitle", parent=styles["Heading1"], fontSize=12, alignment=1, spaceAfter=4,
            )
            subtitle_style = ParagraphStyle(
                "GARSubtitle", parent=styles["Normal"], fontSize=10, alignment=1, spaceAfter=2,
            )
            cell_style = ParagraphStyle(
                "CellStyle", parent=styles["Normal"], fontSize=8, leading=10,
            )
            result_style = ParagraphStyle(
                "ResultStyle", parent=styles["Normal"], fontSize=7, leading=9, alignment=1,
            )

            # === HEADER: Logos + Government Letterhead ===
            logo_size = 45
            logo_row = []
            if dilg_logo:
                logo_row.append(Image(dilg_logo, width=logo_size, height=logo_size))
            else:
                logo_row.append("")
            if bp_logo:
                logo_row.append(Image(bp_logo, width=logo_size, height=logo_size))
            else:
                logo_row.append("")

            if any(logo_row):
                logo_table = Table([logo_row], colWidths=[logo_size + 10, logo_size + 10])
                logo_table.setStyle(TableStyle([
                    ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ]))
                # Wrap in outer table to center the logo pair
                outer = Table([[logo_table]], colWidths=[page_w - 1.0 * inch])
                outer.setStyle(TableStyle([("ALIGN", (0, 0), (0, 0), "CENTER")]))
                elements.append(outer)
                elements.append(Spacer(1, 6))

            # Government letterhead text
            elements.append(Paragraph("Republic of the Philippines", header_text_style))
            elements.append(Paragraph("<b>DEPARTMENT OF THE INTERIOR AND LOCAL GOVERNMENT</b>", header_bold_style))
            elements.append(Paragraph("Region XI \u2013 Province of Davao del Sur", header_text_style))
            elements.append(Paragraph("Office of the Municipal Local Government Operations Officer", header_text_style))
            elements.append(Paragraph("Municipal Local Governance Resource and peace Center", header_text_style))
            elements.append(Paragraph("Municipality of Sulop", header_text_style))
            elements.append(Spacer(1, 4))
            elements.append(Paragraph("<i>\u201cHiniusang Pangatungdanan\u201d</i>", ParagraphStyle(
                "Tagline", parent=styles["Normal"], fontSize=10, alignment=1, fontName="Helvetica-Oblique",
            )))
            elements.append(Spacer(1, 12))

            # === REPORT TITLE ===
            elements.append(Paragraph(gar_data.cycle_year, title_style))
            elements.append(Paragraph("BARANGAY GOVERNANCE ASSESSMENT REPORT", title_style))
            elements.append(Paragraph(
                f"<i>Barangay {gar_data.barangay_name}, {gar_data.municipality}, {gar_data.province}</i>",
                subtitle_style,
            ))
            elements.append(Spacer(1, 14))

            # Color definitions
            met_color = colors.Color(0.56, 0.93, 0.56)
            considered_color = colors.Color(1, 1, 0)
            unmet_color = colors.Color(1, 0.42, 0.42)
            header_color = colors.Color(0.56, 0.93, 0.56)

            # Available content width for portrait A4
            content_width = page_w - 1.0 * inch
            indicator_col_w = content_width - 1.2 * inch
            result_col_w = 1.2 * inch

            # === GOVERNANCE AREAS ===
            for area_idx, area in enumerate(gar_data.governance_areas):
                area_prefix = "CGA" if area.area_type == "Core" else "EGA"
                area_num = area.area_number if area.area_type == "Core" else area.area_number - 3
                area_title = f"{area_prefix} {area_num}: {area.area_name.upper()}"

                table_data = []
                # Area header row
                table_data.append([Paragraph(f"<b>{area_title}</b>", cell_style), ""])
                # Column headers with result legend
                result_header = (
                    '<b>RESULT</b><br/>'
                    '<font size="6">(<font color="green">met</font>, '
                    '<font color="#CC8800">considered</font>, '
                    '<font color="red">unmet</font>)</font>'
                )
                table_data.append([
                    Paragraph("<b>INDICATORS</b>", cell_style),
                    Paragraph(result_header, result_style),
                ])

                style_commands = [
                    ("SPAN", (0, 0), (1, 0)),
                    ("BACKGROUND", (0, 0), (1, 0), header_color),
                    ("ALIGN", (0, 0), (1, 0), "CENTER"),
                    ("BACKGROUND", (0, 1), (1, 1), colors.lightgrey),
                    ("ALIGN", (0, 1), (1, 1), "CENTER"),
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 6),
                    ("TOPPADDING", (0, 0), (-1, -1), 5),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 5),
                    ("ALIGN", (1, 2), (1, -1), "CENTER"),
                ]

                row_idx = 2

                for indicator in area.indicators:
                    indent = "&nbsp;&nbsp;" * indicator.indent_level
                    indicator_text = f"{indent}<b>{indicator.indicator_code}</b> {indicator.indicator_name}"

                    # Result text for non-header indicators
                    result_text = ""
                    if not indicator.is_header and indicator.validation_status:
                        status_upper = indicator.validation_status.upper()
                        if status_upper == "PASS":
                            result_text = "Met"
                            style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), met_color))
                        elif status_upper == "CONDITIONAL":
                            result_text = "Cons."
                            style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), considered_color))
                        elif status_upper == "FAIL":
                            result_text = "Unmet"
                            style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), unmet_color))

                    result_cell = Paragraph(f"<b>{result_text}</b>", result_style) if result_text else ""
                    table_data.append([Paragraph(indicator_text, cell_style), result_cell])

                    if indicator.is_header:
                        style_commands.append(("BACKGROUND", (0, row_idx), (0, row_idx), colors.Color(0.9, 0.97, 0.9)))

                    row_idx += 1

                    # Checklist items
                    for item in indicator.checklist_items:
                        item_text = f"&nbsp;&nbsp;&nbsp;&nbsp;{item.label}"
                        checklist_result = ""
                        if item.validation_result:
                            if item.validation_result == "met":
                                checklist_result = "Met"
                                style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), met_color))
                            elif item.validation_result == "considered":
                                checklist_result = "Cons."
                                style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), considered_color))
                            elif item.validation_result == "unmet":
                                checklist_result = "Unmet"
                                style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), unmet_color))

                        result_cell = Paragraph(checklist_result, result_style) if checklist_result else ""
                        table_data.append([Paragraph(item_text, cell_style), result_cell])
                        row_idx += 1

                # Overall result row
                overall_text = ""
                if area.overall_result:
                    overall_text = area.overall_result.upper()
                    result_color = met_color if area.overall_result == "Passed" else unmet_color
                    style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), result_color))
                style_commands.append(("BACKGROUND", (0, row_idx), (0, row_idx), header_color))
                table_data.append([
                    Paragraph("<b>OVERALL RESULT</b>", cell_style),
                    Paragraph(f"<b>{overall_text}</b>", result_style),
                ])

                col_widths = [indicator_col_w, result_col_w]
                table = Table(table_data, colWidths=col_widths)
                table.setStyle(TableStyle(style_commands))
                elements.append(table)

                # Page break between areas (not after last)
                if area_idx < len(gar_data.governance_areas) - 1:
                    elements.append(PageBreak())
                else:
                    elements.append(Spacer(1, 20))

            # === SUMMARY TABLE ===
            if len(gar_data.governance_areas) > 1:
                elements.append(Spacer(1, 10))

                summary_data = []
                summary_style_commands = [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]

                summary_data.append([
                    Paragraph("<b>Core Governance Area</b>", cell_style),
                    Paragraph("<b>Result</b>", cell_style),
                ])
                summary_style_commands.append(("BACKGROUND", (0, 0), (1, 0), colors.lightgrey))

                row_idx = 1
                for item in gar_data.summary:
                    if item.area_type == "Core":
                        result_text = item.result.upper() if item.result else ""
                        summary_data.append([
                            Paragraph(item.area_name, cell_style),
                            Paragraph(f"<b>{result_text}</b>", result_style),
                        ])
                        if item.result:
                            rc = met_color if item.result == "Passed" else unmet_color
                            summary_style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), rc))
                        row_idx += 1

                summary_data.append([
                    Paragraph("<b>Essential Governance Area</b>", cell_style),
                    Paragraph("<b>Result</b>", cell_style),
                ])
                summary_style_commands.append(("BACKGROUND", (0, row_idx), (1, row_idx), colors.lightgrey))
                row_idx += 1

                for item in gar_data.summary:
                    if item.area_type == "Essential":
                        result_text = item.result.upper() if item.result else ""
                        summary_data.append([
                            Paragraph(item.area_name, cell_style),
                            Paragraph(f"<b>{result_text}</b>", result_style),
                        ])
                        if item.result:
                            rc = met_color if item.result == "Passed" else unmet_color
                            summary_style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), rc))
                        row_idx += 1

                summary_col_widths = [content_width - 1.5 * inch, 1.5 * inch]
                summary_table = Table(summary_data, colWidths=summary_col_widths)
                summary_table.setStyle(TableStyle(summary_style_commands))
                elements.append(summary_table)

            # === BBI COMPLIANCE ===
            if gar_data.bbi_compliance and gar_data.bbi_compliance.bbi_results:
                elements.append(Spacer(1, 30))

                bbi_highly_color = colors.Color(0.56, 0.93, 0.56)
                bbi_mod_color = colors.Color(1, 0.84, 0)
                bbi_low_color = colors.Color(1, 0.42, 0.42)
                bbi_header_color = colors.Color(0.69, 0.77, 0.87)

                bbi_table_data = []
                bbi_style_commands = [
                    ("GRID", (0, 0), (-1, -1), 0.5, colors.black),
                    ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                    ("LEFTPADDING", (0, 0), (-1, -1), 8),
                    ("RIGHTPADDING", (0, 0), (-1, -1), 8),
                    ("TOPPADDING", (0, 0), (-1, -1), 6),
                    ("BOTTOMPADDING", (0, 0), (-1, -1), 6),
                ]

                bbi_table_data.append([
                    Paragraph("<b>BBI FUNCTIONALITY COMPLIANCE (DILG MC 2024-417)</b>", cell_style),
                    "",
                ])
                bbi_style_commands.append(("SPAN", (0, 0), (1, 0)))
                bbi_style_commands.append(("BACKGROUND", (0, 0), (1, 0), bbi_header_color))
                bbi_style_commands.append(("ALIGN", (0, 0), (1, 0), "CENTER"))

                bbi_table_data.append([
                    Paragraph("<b>BBI</b>", cell_style),
                    Paragraph("<b>COMPLIANCE</b>", cell_style),
                ])
                bbi_style_commands.append(("BACKGROUND", (0, 1), (1, 1), colors.lightgrey))
                bbi_style_commands.append(("ALIGN", (0, 1), (1, 1), "CENTER"))

                row_idx = 2
                for bbi in gar_data.bbi_compliance.bbi_results:
                    rating = bbi.compliance_rating.upper() if bbi.compliance_rating else ""
                    rating_label = ""
                    if "HIGHLY" in rating:
                        rating_label = "Highly"
                        bbi_style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), bbi_highly_color))
                    elif "MODERATELY" in rating:
                        rating_label = "Moderate"
                        bbi_style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), bbi_mod_color))
                    elif "LOW" in rating:
                        rating_label = "Low"
                        bbi_style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), bbi_low_color))

                    bbi_table_data.append([
                        Paragraph(f"{bbi.bbi_abbreviation} - {bbi.bbi_name}", cell_style),
                        Paragraph(f"<b>{round(bbi.compliance_percentage)}%</b> ({rating_label})", result_style),
                    ])
                    bbi_style_commands.append(("ALIGN", (1, row_idx), (1, row_idx), "CENTER"))
                    row_idx += 1

                avg_pct = gar_data.bbi_compliance.summary.average_compliance_percentage
                bbi_table_data.append([
                    Paragraph("<b>AVERAGE COMPLIANCE</b>", cell_style),
                    Paragraph(f"<b>{round(avg_pct)}%</b>", result_style),
                ])
                bbi_style_commands.append(("BACKGROUND", (0, row_idx), (0, row_idx), bbi_header_color))
                bbi_style_commands.append(("ALIGN", (1, row_idx), (1, row_idx), "CENTER"))
                if avg_pct >= 75:
                    bbi_style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), bbi_highly_color))
                elif avg_pct >= 50:
                    bbi_style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), bbi_mod_color))
                else:
                    bbi_style_commands.append(("BACKGROUND", (1, row_idx), (1, row_idx), bbi_low_color))

                bbi_col_widths = [content_width - 1.5 * inch, 1.5 * inch]
                bbi_table = Table(bbi_table_data, colWidths=bbi_col_widths)
                bbi_table.setStyle(TableStyle(bbi_style_commands))
                elements.append(bbi_table)

                legend_style = ParagraphStyle(
                    "BBILegend", parent=styles["Normal"], fontSize=8, spaceAfter=4,
                )
                elements.append(Spacer(1, 8))
                summary_text = (
                    f"Summary: {gar_data.bbi_compliance.summary.highly_functional_count} Highly Functional, "
                    f"{gar_data.bbi_compliance.summary.moderately_functional_count} Moderately Functional, "
                    f"{gar_data.bbi_compliance.summary.low_functional_count} Low Functional"
                )
                elements.append(Paragraph(summary_text, legend_style))
                elements.append(Paragraph(
                    "Rating: 75%+ = Highly Functional, 50-74% = Moderately Functional, &lt;50% = Low Functional",
                    legend_style,
                ))

            # Build PDF with footer on every page
            doc.build(elements, onFirstPage=_draw_footer, onLaterPages=_draw_footer)
            buffer.seek(0)

            return buffer

        except ImportError:
            self.logger.error("reportlab not installed. PDF export unavailable.")
            raise ImportError("PDF export requires reportlab. Install with: pip install reportlab")


# Singleton instance
gar_export_service = GARExportService()
