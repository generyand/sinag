# 📊 Municipal Export Service
# Business logic for generating comprehensive municipal data exports

import io
import logging
from datetime import datetime
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import func
from sqlalchemy.orm import Session, joinedload

from app.db.enums import AssessmentStatus, ComplianceStatus, UserRole
from app.db.models.admin import AssessmentCycle
from app.db.models.assessment import Assessment, AssessmentResponse
from app.db.models.barangay import Barangay
from app.db.models.governance_area import GovernanceArea, Indicator
from app.db.models.user import User
from app.schemas.municipal_export import ExportOptions

logger = logging.getLogger(__name__)

# Excel styling constants
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
HEADER_FILL = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
SUBHEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
WRAP_ALIGN = Alignment(horizontal="left", vertical="center", wrap_text=True)


class MunicipalExportService:
    """Service for generating comprehensive municipal data exports."""

    def get_available_cycles(self, db: Session) -> list[dict[str, Any]]:
        """Get list of available assessment cycles for export."""
        cycles = db.query(AssessmentCycle).order_by(AssessmentCycle.year.desc()).all()

        active_cycle_id = None
        result = []
        for cycle in cycles:
            if cycle.is_active:
                active_cycle_id = cycle.id
            result.append(
                {
                    "id": cycle.id,
                    "name": cycle.name,
                    "year": cycle.year,
                    "is_active": cycle.is_active,
                }
            )

        return result, active_cycle_id

    def get_export_data_types(self) -> list[dict[str, Any]]:
        """Get list of available export data types."""
        return [
            {
                "key": "assessments",
                "label": "Assessment Submissions",
                "description": "All assessment submissions with status, scores, and compliance",
                "default": True,
            },
            {
                "key": "analytics",
                "label": "Analytics Summary",
                "description": "Compliance rates, pass/fail statistics, and trends",
                "default": True,
            },
            {
                "key": "governance_areas",
                "label": "Governance Area Performance",
                "description": "Performance breakdown by governance area",
                "default": True,
            },
            {
                "key": "indicators",
                "label": "Indicator Details",
                "description": "Detailed indicator-level breakdown per barangay",
                "default": False,
            },
            {
                "key": "users",
                "label": "BLGU Users",
                "description": "List of registered BLGU users and their barangays",
                "default": False,
            },
        ]

    def generate_export(
        self,
        db: Session,
        options: ExportOptions,
    ) -> tuple[io.BytesIO, dict[str, Any]]:
        """
        Generate comprehensive municipal export as Excel file.

        Args:
            db: Database session
            options: Export options specifying what to include

        Returns:
            tuple: (BytesIO buffer with Excel file, summary metadata)
        """
        wb = Workbook()

        # Remove default sheet
        wb.remove(wb.active)

        cycle = None
        if options.cycle_id:
            cycle = db.query(AssessmentCycle).filter(AssessmentCycle.id == options.cycle_id).first()

        summary = {
            "generated_at": datetime.utcnow(),
            "cycle_name": cycle.name if cycle else "All Cycles",
            "cycle_year": cycle.year if cycle else None,
            "total_barangays": 0,
            "total_assessments": 0,
            "included_sections": [],
        }

        # 1. Summary Sheet (always included)
        self._create_summary_sheet(wb, db, options, cycle, summary)

        # Get the year from cycle for filtering
        filter_year = cycle.year if cycle else None

        # 2. Assessment Submissions
        if options.include_assessments:
            self._create_assessments_sheet(wb, db, filter_year, summary)
            summary["included_sections"].append("Assessments")

        # 3. Analytics
        if options.include_analytics:
            self._create_analytics_sheet(wb, db, filter_year, summary)
            summary["included_sections"].append("Analytics")

        # 4. Governance Area Performance
        if options.include_governance_areas:
            self._create_governance_areas_sheet(wb, db, filter_year)
            summary["included_sections"].append("Governance Areas")

        # 5. Indicator Details
        if options.include_indicators:
            self._create_indicators_sheet(wb, db, filter_year)
            summary["included_sections"].append("Indicators")

        # 6. Users
        if options.include_users:
            self._create_users_sheet(wb, db)
            summary["included_sections"].append("Users")

        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)

        return buffer, summary

    def _create_summary_sheet(
        self,
        wb: Workbook,
        db: Session,
        options: ExportOptions,
        cycle: AssessmentCycle | None,
        summary: dict,
    ) -> None:
        """Create summary/metadata sheet."""
        ws = wb.create_sheet("Summary", 0)

        # Title
        ws.merge_cells("A1:D1")
        title_cell = ws["A1"]
        title_cell.value = "SINAG Municipal Data Export"
        title_cell.font = Font(bold=True, size=16, color="003366")
        title_cell.alignment = CENTER_ALIGN

        # Metadata
        metadata = [
            ("Generated At", summary["generated_at"].strftime("%Y-%m-%d %H:%M:%S UTC")),
            ("Assessment Cycle", summary["cycle_name"]),
            ("Cycle Year", summary["cycle_year"] or "All"),
            ("", ""),
            ("Total Barangays", db.query(func.count(Barangay.id)).scalar() or 0),
        ]

        # Count assessments
        assessment_query = db.query(func.count(Assessment.id))
        if cycle:
            assessment_query = assessment_query.filter(Assessment.assessment_year == cycle.year)
        total_assessments = assessment_query.scalar() or 0
        metadata.append(("Total Assessments", total_assessments))

        summary["total_barangays"] = metadata[4][1]
        summary["total_assessments"] = total_assessments

        # Write metadata
        for row_num, (label, value) in enumerate(metadata, 3):
            ws.cell(row=row_num, column=1, value=label).font = Font(bold=True)
            ws.cell(row=row_num, column=2, value=value)

        # Included sections
        row = len(metadata) + 4
        ws.cell(row=row, column=1, value="Included Sections:").font = Font(bold=True)
        row += 1
        sections = []
        if options.include_assessments:
            sections.append("Assessment Submissions")
        if options.include_analytics:
            sections.append("Analytics Summary")
        if options.include_governance_areas:
            sections.append("Governance Area Performance")
        if options.include_indicators:
            sections.append("Indicator Details")
        if options.include_users:
            sections.append("BLGU Users")

        for section in sections:
            ws.cell(row=row, column=1, value=f"  • {section}")
            row += 1

        # Adjust column widths
        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 40

    def _create_assessments_sheet(
        self,
        wb: Workbook,
        db: Session,
        year: int | None,
        summary: dict,
    ) -> None:
        """Create assessments sheet with all submissions."""
        ws = wb.create_sheet("Assessments")

        # Headers
        headers = [
            "Barangay",
            "Status",
            "Submitted At",
            "Validated At",
            "Approved At",
            "Compliance Status",
            "Pass Count",
            "Fail Count",
            "Conditional Count",
            "Total Indicators",
            "Score (%)",
            "Gov. Areas Passed",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

        # Query assessments with related data
        query = db.query(Assessment).options(
            joinedload(Assessment.blgu_user).joinedload(User.barangay),
            joinedload(Assessment.responses),
        )
        if year:
            query = query.filter(Assessment.assessment_year == year)

        assessments = query.order_by(Assessment.id).all()

        # Write data rows
        for row_num, assessment in enumerate(assessments, 2):
            barangay_name = "Unknown"
            if assessment.blgu_user and assessment.blgu_user.barangay:
                barangay_name = assessment.blgu_user.barangay.name

            # Count responses
            pass_count = fail_count = conditional_count = 0
            for resp in assessment.responses:
                if resp.validation_status:
                    status_val = (
                        resp.validation_status.value
                        if hasattr(resp.validation_status, "value")
                        else resp.validation_status
                    )
                    if status_val.upper() == "PASS":
                        pass_count += 1
                    elif status_val.upper() == "FAIL":
                        fail_count += 1
                    elif status_val.upper() == "CONDITIONAL":
                        conditional_count += 1

            total = len(assessment.responses)
            score = round(((pass_count + conditional_count) / total * 100), 1) if total > 0 else 0

            row_data = [
                barangay_name,
                assessment.status.value if assessment.status else "",
                assessment.submitted_at.strftime("%Y-%m-%d %H:%M")
                if assessment.submitted_at
                else "",
                assessment.validated_at.strftime("%Y-%m-%d %H:%M")
                if assessment.validated_at
                else "",
                assessment.mlgoo_approved_at.strftime("%Y-%m-%d %H:%M")
                if assessment.mlgoo_approved_at
                else "",
                assessment.final_compliance_status.value
                if assessment.final_compliance_status
                else "",
                pass_count,
                fail_count,
                conditional_count,
                total,
                score,
                self._count_passed_areas(assessment),
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN if col > 1 else WRAP_ALIGN

        # Adjust column widths
        column_widths = [25, 20, 18, 18, 18, 15, 10, 10, 12, 12, 10, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # Freeze header
        ws.freeze_panes = "A2"

    def _count_passed_areas(self, assessment: Assessment) -> int:
        """Count how many governance areas passed (≥70% pass rate)."""
        areas_data = {}
        for resp in assessment.responses:
            if not resp.indicator or not resp.indicator.governance_area_id:
                continue
            area_id = resp.indicator.governance_area_id
            if area_id not in areas_data:
                areas_data[area_id] = {"pass": 0, "total": 0}
            areas_data[area_id]["total"] += 1
            if resp.validation_status:
                status_val = (
                    resp.validation_status.value
                    if hasattr(resp.validation_status, "value")
                    else resp.validation_status
                )
                if status_val.upper() in ["PASS", "CONDITIONAL"]:
                    areas_data[area_id]["pass"] += 1

        passed = 0
        for stats in areas_data.values():
            if stats["total"] > 0 and (stats["pass"] / stats["total"]) >= 0.7:
                passed += 1
        return passed

    def _create_analytics_sheet(
        self,
        wb: Workbook,
        db: Session,
        year: int | None,
        summary: dict,
    ) -> None:
        """Create analytics summary sheet."""
        ws = wb.create_sheet("Analytics")

        # Title
        ws.merge_cells("A1:C1")
        ws["A1"].value = "Analytics Summary"
        ws["A1"].font = Font(bold=True, size=14)

        # Query data
        base_query = db.query(Assessment)
        if year:
            base_query = base_query.filter(Assessment.assessment_year == year)

        total_assessments = base_query.count()
        completed = base_query.filter(Assessment.status == AssessmentStatus.COMPLETED).count()
        passed = base_query.filter(
            Assessment.status == AssessmentStatus.COMPLETED,
            Assessment.final_compliance_status == ComplianceStatus.PASSED,
        ).count()
        failed = base_query.filter(
            Assessment.status == AssessmentStatus.COMPLETED,
            Assessment.final_compliance_status == ComplianceStatus.FAILED,
        ).count()

        # Write analytics
        analytics = [
            ("", ""),
            ("Compliance Statistics", ""),
            ("Total Assessments", total_assessments),
            ("Completed Assessments", completed),
            ("Passed (Compliant)", passed),
            ("Failed (Non-Compliant)", failed),
            ("Compliance Rate", f"{(passed / completed * 100):.1f}%" if completed > 0 else "N/A"),
            ("", ""),
            ("Status Distribution", ""),
        ]

        # Add status distribution
        for status in AssessmentStatus:
            count = base_query.filter(Assessment.status == status).count()
            if count > 0:
                analytics.append((status.value, count))

        for row_num, (label, value) in enumerate(analytics, 3):
            label_cell = ws.cell(row=row_num, column=1, value=label)
            ws.cell(row=row_num, column=2, value=value)

            if label and not value:  # Section header
                label_cell.font = Font(bold=True, size=12)
            elif label:
                label_cell.font = Font(bold=True)

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 20

    def _create_governance_areas_sheet(
        self,
        wb: Workbook,
        db: Session,
        year: int | None,
    ) -> None:
        """Create governance area performance sheet."""
        ws = wb.create_sheet("Governance Areas")

        headers = [
            "Governance Area",
            "Code",
            "Type",
            "Total Indicators",
            "Avg Pass Rate (%)",
            "Best Performing Barangay",
            "Lowest Performing Barangay",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        # Get governance areas
        areas = db.query(GovernanceArea).order_by(GovernanceArea.code).all()

        row_num = 2
        for area in areas:
            indicator_count = (
                db.query(func.count(Indicator.id))
                .filter(
                    Indicator.governance_area_id == area.id,
                    Indicator.is_active == True,
                )
                .scalar()
                or 0
            )

            row_data = [
                area.name,
                area.code or "",
                area.area_type.value if area.area_type else "",
                indicator_count,
                "N/A",  # Would need more complex calculation
                "N/A",
                "N/A",
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = THIN_BORDER

            row_num += 1

        # Adjust widths
        column_widths = [40, 10, 12, 15, 15, 25, 25]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"

    def _create_indicators_sheet(
        self,
        wb: Workbook,
        db: Session,
        year: int | None,
    ) -> None:
        """Create indicator details sheet."""
        ws = wb.create_sheet("Indicators")

        headers = [
            "Barangay",
            "Governance Area",
            "Indicator Code",
            "Indicator Name",
            "Status",
            "Requires Rework",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        # Query responses with details
        query = (
            db.query(AssessmentResponse)
            .join(Assessment)
            .options(
                joinedload(AssessmentResponse.indicator).joinedload(Indicator.governance_area),
                joinedload(AssessmentResponse.assessment)
                .joinedload(Assessment.blgu_user)
                .joinedload(User.barangay),
            )
        )
        if year:
            query = query.filter(Assessment.assessment_year == year)

        responses = query.order_by(Assessment.id, AssessmentResponse.indicator_id).limit(5000).all()

        for row_num, resp in enumerate(responses, 2):
            barangay_name = "Unknown"
            if resp.assessment.blgu_user and resp.assessment.blgu_user.barangay:
                barangay_name = resp.assessment.blgu_user.barangay.name

            area_name = (
                resp.indicator.governance_area.name
                if resp.indicator and resp.indicator.governance_area
                else ""
            )
            indicator_code = resp.indicator.indicator_code if resp.indicator else ""
            indicator_name = resp.indicator.name if resp.indicator else ""
            status = resp.validation_status.value if resp.validation_status else "NOT_VALIDATED"

            row_data = [
                barangay_name,
                area_name,
                indicator_code,
                indicator_name,
                status,
                "Yes" if resp.requires_rework else "No",
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = THIN_BORDER

        # Adjust widths
        column_widths = [25, 35, 12, 50, 15, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"

    def _create_users_sheet(
        self,
        wb: Workbook,
        db: Session,
    ) -> None:
        """Create BLGU users sheet."""
        ws = wb.create_sheet("Users")

        headers = [
            "Name",
            "Email",
            "Role",
            "Barangay",
            "Is Active",
            "Created At",
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER

        # Query BLGU users
        users = (
            db.query(User)
            .options(joinedload(User.barangay))
            .filter(User.role == UserRole.BLGU_USER)
            .order_by(User.name)
            .all()
        )

        for row_num, user in enumerate(users, 2):
            row_data = [
                user.name or "",
                user.email,
                user.role.value if user.role else "",
                user.barangay.name if user.barangay else "",
                "Yes" if user.is_active else "No",
                user.created_at.strftime("%Y-%m-%d") if user.created_at else "",
            ]

            for col, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_num, column=col, value=value)
                cell.border = THIN_BORDER

        # Adjust widths
        column_widths = [25, 35, 15, 25, 10, 15]
        for col, width in enumerate(column_widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        ws.freeze_panes = "A2"


# Singleton instance
municipal_export_service = MunicipalExportService()
