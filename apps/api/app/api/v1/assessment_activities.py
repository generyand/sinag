# 📊 Assessment Activities API Routes
# Endpoints for retrieving assessment workflow activity logs

import io
from datetime import datetime

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

from app.api.deps import get_db, require_mlgoo_dilg
from app.db.models.user import User
from app.schemas.assessment_activity import (
    ActivityByActionCount,
    ActivityCountsResponse,
    ActivitySummary,
    AssessmentActivityListResponse,
    AssessmentActivityResponse,
    AssessmentTimelineItem,
    AssessmentTimelineResponse,
)
from app.services.assessment_activity_service import assessment_activity_service

router = APIRouter()


# ============================================================================
# Assessment Activity Endpoints
# ============================================================================


@router.get(
    "/",
    response_model=AssessmentActivityListResponse,
    tags=["assessment-activities"],
    summary="Get assessment activities with filtering",
    description="Retrieve assessment workflow activities with optional filtering. Requires MLGOO_DILG role.",
)
async def get_activities(
    skip: int = Query(0, ge=0, description="Number of records to skip"),
    limit: int = Query(100, ge=1, le=500, description="Maximum records to return"),
    assessment_id: int | None = Query(None, description="Filter by assessment ID"),
    user_id: int | None = Query(None, description="Filter by user ID"),
    barangay_id: int | None = Query(None, description="Filter by barangay ID"),
    action: str | None = Query(None, description="Filter by single action type"),
    actions: list[str] | None = Query(None, description="Filter by multiple action types"),
    start_date: datetime | None = Query(None, description="Filter from date"),
    end_date: datetime | None = Query(None, description="Filter to date"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_mlgoo_dilg),
):
    """
    Get assessment activities with optional filtering and pagination.

    **Authentication:** Requires MLGOO_DILG role.

    **Filters:**
    - `assessment_id`: Filter by specific assessment
    - `user_id`: Filter by user who performed the action
    - `barangay_id`: Filter by barangay
    - `action`: Filter by single action type (submitted, approved, etc.)
    - `actions`: Filter by multiple action types (e.g., actions=submitted&actions=approved)
    - `start_date`, `end_date`: Filter by date range

    **Returns:**
    - Paginated list of activities with user and assessment details
    """
    activities, total = assessment_activity_service.get_activities(
        db=db,
        skip=skip,
        limit=limit,
        assessment_id=assessment_id,
        user_id=user_id,
        barangay_id=barangay_id,
        action=action,
        actions=actions,
        start_date=start_date,
        end_date=end_date,
    )

    # Enrich with user and assessment details
    items = []
    for activity in activities:
        barangay_name = None
        assessment_year = None
        if activity.assessment:
            if activity.assessment.blgu_user and activity.assessment.blgu_user.barangay:
                barangay_name = activity.assessment.blgu_user.barangay.name
            # assessment_year is an integer (the year number), not a relationship
            assessment_year = activity.assessment.assessment_year

        items.append(
            AssessmentActivityResponse(
                id=activity.id,
                assessment_id=activity.assessment_id,
                user_id=activity.user_id,
                action=activity.action,
                from_status=activity.from_status,
                to_status=activity.to_status,
                extra_data=activity.extra_data,
                description=activity.description,
                created_at=activity.created_at,
                user_email=activity.user.email if activity.user else None,
                user_name=activity.user.name if activity.user else None,
                barangay_name=barangay_name,
                assessment_year=assessment_year,
            )
        )

    return AssessmentActivityListResponse(
        items=items,
        total=total,
        skip=skip,
        limit=limit,
    )


@router.get(
    "/timeline/{assessment_id}",
    response_model=AssessmentTimelineResponse,
    tags=["assessment-activities"],
    summary="Get assessment timeline",
    description="Get complete activity timeline for a specific assessment. Requires MLGOO_DILG role.",
)
async def get_assessment_timeline(
    assessment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(require_mlgoo_dilg),
):
    """
    Get the complete activity timeline for a specific assessment.

    **Authentication:** Requires MLGOO_DILG role.

    **Returns:**
    - Assessment details with chronological activity timeline
    """
    from fastapi import HTTPException, status

    from app.db.models.assessment import Assessment

    # Get assessment details
    assessment = db.query(Assessment).filter(Assessment.id == assessment_id).first()
    if not assessment:
        raise HTTPException(
            status_code=status.HTTP_404_NOT_FOUND,
            detail=f"Assessment with ID {assessment_id} not found",
        )

    # Get barangay name
    barangay_name = "Unknown Barangay"
    if assessment.blgu_user and assessment.blgu_user.barangay:
        barangay_name = assessment.blgu_user.barangay.name

    # Get timeline
    activities = assessment_activity_service.get_assessment_timeline(db, assessment_id)

    timeline_items = []
    for activity in activities:
        timeline_items.append(
            AssessmentTimelineItem(
                id=activity.id,
                action=activity.action,
                from_status=activity.from_status,
                to_status=activity.to_status,
                description=activity.description,
                extra_data=activity.extra_data,
                created_at=activity.created_at,
                user_id=activity.user_id,
                user_email=activity.user.email if activity.user else None,
                user_name=activity.user.name if activity.user else None,
                user_role=activity.user.role.value if activity.user else None,
            )
        )

    return AssessmentTimelineResponse(
        assessment_id=assessment_id,
        barangay_name=barangay_name,
        current_status=assessment.status.value,
        timeline=timeline_items,
    )


@router.get(
    "/summary",
    response_model=ActivityCountsResponse,
    tags=["assessment-activities"],
    summary="Get activity summary statistics",
    description="Get summary statistics of assessment activities. Requires MLGOO_DILG role.",
)
async def get_activity_summary(
    barangay_id: int | None = Query(None, description="Filter by barangay ID"),
    start_date: datetime | None = Query(None, description="Filter from date"),
    end_date: datetime | None = Query(None, description="Filter to date"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_mlgoo_dilg),
):
    """
    Get activity summary statistics.

    **Authentication:** Requires MLGOO_DILG role.

    **Returns:**
    - Summary statistics (total, submissions, approvals, etc.)
    - Counts by action type
    """
    summary = assessment_activity_service.get_activity_summary(db, barangay_id)
    by_action = assessment_activity_service.get_activity_counts_by_action(
        db, barangay_id, start_date, end_date
    )

    return ActivityCountsResponse(
        summary=ActivitySummary(**summary),
        by_action=[ActivityByActionCount(**item) for item in by_action],
    )


@router.get(
    "/export",
    tags=["assessment-activities"],
    summary="Export activities to Excel",
    description="Export assessment activities to Excel format. Requires MLGOO_DILG role.",
)
async def export_activities_excel(
    assessment_id: int | None = Query(None, description="Filter by assessment ID"),
    user_id: int | None = Query(None, description="Filter by user ID"),
    barangay_id: int | None = Query(None, description="Filter by barangay ID"),
    action: str | None = Query(None, description="Filter by action type"),
    start_date: datetime | None = Query(None, description="Filter from date"),
    end_date: datetime | None = Query(None, description="Filter to date"),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_mlgoo_dilg),
):
    """
    Export assessment activities to Excel with optional filtering.

    **Authentication:** Requires MLGOO_DILG role.

    **Returns:**
    - Excel file (.xlsx) with activity data
    """
    # Get all matching activities (high limit for export)
    activities, _ = assessment_activity_service.get_activities(
        db=db,
        skip=0,
        limit=10000,
        assessment_id=assessment_id,
        user_id=user_id,
        barangay_id=barangay_id,
        action=action,
        start_date=start_date,
        end_date=end_date,
    )

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Activity Logs"

    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    # Headers
    headers = [
        "ID",
        "Date/Time",
        "Action",
        "Description",
        "Barangay",
        "User",
        "User Email",
        "Previous Status",
        "New Status",
        "Assessment ID",
    ]

    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border

    # Data rows
    for row_num, activity in enumerate(activities, 2):
        barangay_name = ""
        if (
            activity.assessment
            and activity.assessment.blgu_user
            and activity.assessment.blgu_user.barangay
        ):
            barangay_name = activity.assessment.blgu_user.barangay.name

        row_data = [
            activity.id,
            activity.created_at.strftime("%Y-%m-%d %H:%M:%S") if activity.created_at else "",
            activity.action,
            activity.description or "",
            barangay_name,
            activity.user.name if activity.user else "",
            activity.user.email if activity.user else "",
            activity.from_status or "",
            activity.to_status or "",
            activity.assessment_id,
        ]

        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")

    # Adjust column widths
    column_widths = [8, 20, 25, 35, 25, 20, 30, 20, 20, 12]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width

    # Freeze header row
    ws.freeze_panes = "A2"

    # Save to buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"activity_logs_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        buffer,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )
